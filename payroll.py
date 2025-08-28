from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
import re
from openpyxl import load_workbook
import os
import sys
from io import BytesIO
from rapidfuzz import process, fuzz
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment
from copy import copy
import difflib
from datetime import timedelta
# --- NEW IMPORTS for browser opening ---
import threading
import webbrowser
import logging # Good practice for logging errors
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
from datetime import datetime
from werkzeug.utils import secure_filename
CPT_REGEX = re.compile(r'^993\d{2}$')
# Global variables (declared here for scope)
provider_cpt_dict = {}
not_recognized = {}
current_workbook = None
output_filename = ""
sheet = None
payroll_df = None
common_providers = []
practitioner_list = []
cpt_positions = {}
start_row = None # Add these to your globals
row_end = None   # Add these to your globals
merged_cells = None # Add this to your globals
capture_df = None # Add this to your globals
missing_providers=[]
cpt_counts=pd.DataFrame()
Gross_encounters_col=None
week1_encounters_col_idx=None
week2_encounters_col_idx=None
# Add this to your global variable declarations
manual_cpt_updates = []
weekly_counts=pd.DataFrame()
processing_warnings = [] # IMPORTANT: Reset the warnings list here
date_range="" # Store the date range for use in other functions
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS  # PyInstaller extracts to this temp dir
    return os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__,
            template_folder=os.path.join(get_base_dir(), 'templates'))

BASE_DIR = get_base_dir()
print("Base Directory:", BASE_DIR)
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


def remove_cpt_from_providers(data_dict, cpt_to_remove):
    """
    Removes a given CPT code from each provider's dictionary in the main data_dict.

    Args:
        data_dict (dict): The dictionary containing provider data,
                          e.g., {'ProviderName': {'CPTCode1': Count1, 'CPTCode2': Count2}}.
        cpt_to_remove (str): The CPT code string to be removed.

    Returns:
        dict: The updated dictionary with the specified CPT code removed.
    """
    updated_dict = {}
    for provider, cpt_counts in data_dict.items():
        # Create a new dictionary for the current provider's CPTs
        # This prevents modifying the dictionary while iterating
        new_cpt_counts = {cpt: count for cpt, count in cpt_counts.items() if cpt != cpt_to_remove}
        
        # Only add the provider to the updated_dict if they still have CPTs
        if new_cpt_counts:
            updated_dict[provider] = new_cpt_counts
    return updated_dict
def _norm(s: str) -> str:
    s = str(s).lower().strip()
    s = re.sub(r"[^\w\s'-]", " ", s)
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def _parse_name(s: str):
    s = _norm(s)
    parts = s.split()
    if not parts:
        return {"first":"", "middles":[], "last":"", "tokens":set()}
    prefixes = {"dr","mr","mrs","ms","miss"}
    suffixes = {"jr","sr","ii","iii","iv"}
    parts = [p for p in parts if p not in prefixes|suffixes]
    if not parts:
        return {"first":"", "middles":[], "last":"", "tokens":set()}
    first = parts[0]
    last = parts[-1] if len(parts) > 1 else ""
    middles = parts[1:-1] if len(parts) > 2 else []
    return {"first": first, "middles": middles, "last": last, "tokens": set(parts)}

def _first_match_score(a_first: str, b_first: str) -> float:
    if not a_first or not b_first:
        return 0.0
    a_first = a_first.strip().lower()
    b_first = b_first.strip().lower()
    # initial vs full name
    if a_first[0] == b_first[0] and (len(a_first) == 1 or len(b_first) == 1):
        return 95.0
    # fallback similarity
    return float(fuzz.ratio(a_first, b_first))

def _last_match_score(a_last: str, b_last: str) -> float:
    if not a_last or not b_last:
        return 0.0
    a = a_last.replace("'", "")
    b = b_last.replace("'", "")
    return float(max(fuzz.ratio(a, b), fuzz.token_sort_ratio(a, b)))

def name_similarity(a: str, b: str, *, score_cutoff: float = 0.0) -> float:
    pa, pb = _parse_name(a), _parse_name(b)
    last = _last_match_score(pa["last"], pb["last"])
    if last < 75:
        return 0.0
    first = _first_match_score(pa["first"], pb["first"])
    middle = 0.0
    if pa["middles"] and pb["middles"]:
        middle = float(fuzz.token_set_ratio(" ".join(pa["middles"]), " ".join(pb["middles"])))
    full = float(fuzz.token_set_ratio(_norm(a), _norm(b)))
    weighted = 0.7*last + 0.25*first + 0.05*middle
    score = max(weighted, full if last >= 85 else weighted)
    return score if score >= score_cutoff else 0.0

def match_providers(providers, org_providers, threshold=76):
    matched = []
    unmatched_map = {}  # providers_norm -> best unmatched

    matched_providers_set = set()  # to avoid re-matching same provider
    already_matched_org_providers_names = set()

    providers_map = { _norm(p): p for p in providers }
    org_providers_map = { _norm(p): p for p in org_providers }
    comp_norm = list(providers_map.keys())
    cap_norm = list(org_providers_map.keys())

    def last_token(s):
        parts = _norm(s).split()
        return parts[-1] if parts else ""

    # simple blocking
    cap_by_last, cap_by_initial = {}, {}
    for n in cap_norm:
        lt = last_token(n)
        cap_by_last.setdefault(lt, []).append(n)
        cap_by_initial.setdefault(lt[:1], []).append(n)

    for cr in comp_norm:
        original_cr = providers_map[cr]
        if original_cr in matched_providers_set:
            continue

        best_name = None
        best_score = -1.0
        matched_here = False

        for current_threshold in range(90, threshold - 1, -5):
            lt = last_token(cr)
            cand = set(cap_by_last.get(lt, [])) | set(cap_by_initial.get(lt[:1], []))
            if not cand:
                cand = set(cap_norm)
            cand = [c for c in cand if c not in already_matched_org_providers_names]
            if not cand:
                continue

            best = process.extractOne(cr, cand, scorer=name_similarity)
            if best is None:
                continue

            match_name, score, _ = best
            score = float(score)

            if score > best_score:
                best_score = score
                best_name = match_name

            if score >= current_threshold:
                matched.append({
                    'charge_capture_name': original_cr,
                    'payroll_name': org_providers_map[match_name],
                    'score': score
                })
                matched_providers_set.add(original_cr)
                already_matched_org_providers_names.add(match_name)
                matched_here = True
                break

        if not matched_here:
            unmatched_map[cr] = {
                'charge_capture_name': original_cr,
                'payroll_name': org_providers_map[best_name] if best_name and best_score > 0 else None,
                'score': best_score if best_score > 0 else 0.0
            }

    unmatched = sorted(unmatched_map.values(), key=lambda x: x['score'], reverse=True)
    return {"matched": matched, "unmatched": unmatched}
# function for saftey 
def reset_global_variables():
    global provider_cpt_dict, not_recognized, current_workbook, sheet, payroll_df, common_providers, \
           practitioner_list, cpt_positions, start_row, row_end, merged_cells, capture_df, \
           Gross_encounters_col, manual_cpt_updates, weekly_counts, week1_encounters_col_idx, week2_encounters_col_idx, date_range
    provider_cpt_dict = {}
    not_recognized = {}
    current_workbook = None
    sheet = None
    payroll_df = None
    common_providers = []
    practitioner_list = []
    cpt_positions = {}
    start_row = None
    row_end = None
    merged_cells = None
    capture_df = None
    Gross_encounters_col = None
    manual_cpt_updates = []  # Reset it here
    weekly_counts = pd.DataFrame()  # Initialize as empty DataFrame instead of None
    processing_warnings = []
    week1_encounters_col_idx = None
    week2_encounters_col_idx = None
    date_range = ""
# extarct providers from payroll
def extract_practitioners(df, start_row=3, stop_row=102):
    practitioners = []
    for i in range(start_row + 1,stop_row):
        name = df.iloc[i, 2]
        if pd.isna(name):
            continue
        practitioners.append(str(name).strip())
    flipped = [' '.join(n.split(', ')[::-1]) if ', ' in n else n for n in practitioners]
    return flipped

# normalize names to be lowerd and no whitespaces 
def normalize(name):
    try:
        return name.replace(" ", "").lower()
    except:
        return name
# get the cpt in payroll df and get the row and col indices 
def find_header_and_cpt_positions(df):
    # Attempt to find the start and end markers for the CPT code section
    start_match = find_cell_matches(df, "Encounter Pay Model")
    if not start_match:
        raise ValueError("Could not locate 'Encounter Pay Model' header in the payroll file. Please ensure this header exists and is spelled correctly.")

    end_match = find_cell_matches(df, "Week 1 Encounters")
    if not end_match:
        end_match = find_cell_matches(df, "Gross Encounters")
        if not end_match:
            raise ValueError("Could not locate 'Week 1 Encounters' or 'Gross Encounters' header in the payroll file. These headers are needed to define the CPT code section.")

    start_col = start_match[0]['col']
    end_col = end_match[0]['col']
    
    for i, row in df.iterrows():
        cpt_positions = {}
        # Iterate through columns between the determined start and end columns
        for idx in range(start_col + 1, end_col):  # start_col + 1 to exclude the 'Encounter Pay Model' column itself
            value = row.iloc[idx]
            str_val = str(value).strip()

            # Check if the value is not NaN and is a non-empty string
            # Also, add a basic regex check for CPT-like format (e.g., "993xx") to be more robust
            if pd.notna(value) and str_val and re.match(r'^\d{5}$', str_val):  # Basic check for 5 digits for a CPT
                cpt_positions[str_val] = idx

        # If we found at least one potential CPT code in this row, assume it's the header row
        if cpt_positions:
            return i, cpt_positions

    # If the loop finishes without finding a CPT header row
    raise ValueError(
        "Could not find a valid CPT header row in the payroll file within the expected section. "
        "Please ensure CPT codes are present as column headers between 'Encounter Pay Model' "
        "and 'Week 1 Encounters' or 'Gross Encounters'."
    )

# cheack if the current cell is merged with other cells
def is_merged(cell_row, cell_col):
    global merged_cells
    cell_ref = f"{get_column_letter(cell_col)}{cell_row}"
    for merged_range in merged_cells:
        if cell_ref in merged_range:
            return merged_range.min_row, merged_range.min_col
    return cell_row, cell_col

# get cell span (used mainly for getting the table end row )
def get_cell_span(sheet, row, col):
    """
    Given a sheet and cell coordinates (row, col),
    return (row_span, col_span) if merged,
    else (1, 1) if not merged.
    """
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            # merged_range has min_row, max_row, min_col, max_col
            row_span = merged_range.max_row - merged_range.min_row + 1
            col_span = merged_range.max_col - merged_range.min_col + 1
            return row_span, col_span
    return 1, 1  # not merged, single cell

# return cell (row , col ) incices
def find_cell_matches(df, search_text, similarity_threshold=0.95):
    """
    Searches every cell in the DataFrame for fuzzy matches with the given text.
    
    Args:
        df (pd.DataFrame): The DataFrame to search.
        search_text (str): The text to search for.
        similarity_threshold (float): Minimum similarity (0–1) to consider a match.

    Returns:
        list of dicts: Each dict contains row index, column index, column name, and value of the match.
    """
    def normalize(text):
        return str(text).replace('\n', ' ').strip().lower()

    target = normalize(search_text)
    matches = []

    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            cell = df.iat[row_idx, col_idx]
            cell_str = normalize(cell)

            # Compare similarity
            similarity = difflib.SequenceMatcher(None, target, cell_str).ratio()
            if similarity >= similarity_threshold:
                matches.append({
                    'row': row_idx,
                    'col': col_idx,
                    'column_name': df.columns[col_idx],
                    'value': cell,
                    'similarity': round(similarity, 3)
                })
    return matches

# add new cpt col and return new sheet 
def add_new_cpt(cpt_code,Start_col_index):
    global start_row,row_end,sheet
    try:
        # Step 1: Capture and unmerge existing merged cells before inserting
        adjusted_merges = []
        for merged_range in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(merged_range.coord)

            if min_row < row_end and  min_col >= Start_col_index:
                # Adjust the merged range
                adjusted_merges.append([min_row, min_col+5, max_row, max_col+5])
                sheet.unmerge_cells(merged_range.coord)
            if min_col < Start_col_index < max_col and min_row <row_end:
                # Adjust the merged range
                adjusted_merges.append([min_row, min_col, max_row, max_col+5])
                sheet.unmerge_cells(merged_range.coord)
        # Step 2: Insert the column

        sheet.insert_cols(Start_col_index, amount=5)
        # Step 5: Set header value for inserted column

        # Step 3: Copy values and styles from old column to the right
        for row in range(start_row+3, row_end + 1):
            for col in range(Start_col_index,Start_col_index+5):
                old_cell = sheet.cell(row=row, column=col-3)
                new_cell = sheet.cell(row=row, column=col)

                # Copy all styling attributes
                new_cell.font = copy(old_cell.font)
                new_cell.border = copy(old_cell.border)
                new_cell.fill = copy(old_cell.fill)
                new_cell.number_format = copy(old_cell.number_format)
                new_cell.protection = copy(old_cell.protection)
                new_cell.alignment = copy(old_cell.alignment)
        for col in range(Start_col_index,Start_col_index+5):
            old_cell=sheet.cell(row=start_row+5,column=col)
            new_cell=sheet.cell(row=start_row+2, column=col)
                            # Copy all styling attributes
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)
        # Step 4: Reapply adjusted merged cells
        for merge in adjusted_merges:
            min_row, min_col, max_row, max_col = merge
            if min_row != max_row or min_col != max_col:  # Skip single-cell "merges"
                sheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)

        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        cell = sheet.cell(row=start_row + 2, column=Start_col_index)
        cell.value = cpt_code
        cell.font = bold_font
        cell.alignment = center_align
        sheet.merge_cells(start_row=start_row+2,start_column=Start_col_index,end_row=start_row+2,end_column=Start_col_index+4)

        for col_offset, text in enumerate(["Encounters", "Rate", "Sum","Week 1\nEncounters","Week 2\nEncounters"]):
            cell = sheet.cell(row=start_row + 3, column=Start_col_index + col_offset)
            cell.value = text
            cell.font = bold_font
            cell.alignment = center_align
    except Exception as e:
        print("An error occurred:", e)
        return False


#  when add missing cpt code from ui the fucntion modify the data ,(update provider_cpt_dict )
def update_cpt_counts(data: dict, name: str, cpt_code: str) -> None:
    # If name is not in the dictionary, add it with the CPT code count = 1
    if name not in data:
        data[name] = {cpt_code: 1}
    else:
        # If CPT code is not in that person's dict, add it
        if cpt_code not in data[name]:
            data[name][cpt_code] = 1
        else:
            # If CPT code exists, increment the count
            data[name][cpt_code] += 1

# return provider_cpt_dict ( include provider with cpt encounters)
def process_cpt_counts(cpt_counts: pd.DataFrame, cpt_positions) -> None:
    """
    Updates provider_cpt_dict and not_recognized based on the CPT counts DataFrame.

    Parameters:
    - cpt_counts: DataFrame with 'Provider', 'CPT Codes', and 'Counts' columns
    - provider_cpt_dict: Dictionary to store recognized CPT codes per provider
    - not_recognized: Dictionary to store unrecognized CPT codes per provider
    - cpt_pattern: Compiled regex pattern to match valid CPT codes
    """
    global provider_cpt_dict ,not_recognized
    provider_cpt_dict={}
    not_recognized={}
    for _, row in cpt_counts.iterrows():
        provider = row['Provider']
        raw_cpt_code = str(row['CPT Codes'])
        count = row['Counts']

        if provider not in provider_cpt_dict:
            provider_cpt_dict[provider] = {"name": provider}

        cpt_codes = [code.strip() for code in raw_cpt_code.split(",")]
        for cpt in cpt_codes:
            if cpt in cpt_positions:
                if cpt not in provider_cpt_dict[provider]:
                    provider_cpt_dict[provider][cpt] = 0
                provider_cpt_dict[provider][cpt] += count
            else:
                if provider not in not_recognized:
                    not_recognized[provider] = {}
                not_recognized[provider][cpt] = count
                not_recognized[provider]['name'] = provider

#  populate the given sheet with new data and saves it 
def write_provider_cpt_data_to_sheet(payroll_df, common_providers, practitioner_list,
                                     provider_cpt_dict, cpt_positions, output_filename):
    """
    Writes CPT counts and Gross Encounters for matched providers into an Excel sheet.

    Parameters:
    - sheet: The openpyxl worksheet object.
    - payroll_df: The pandas DataFrame of the payroll data.
    - common_providers: List of dicts with 'payroll_name' and 'charge_capture_name'.
    - practitioner_list: List of provider names in the payroll sheet.
    - provider_cpt_dict: Dict mapping provider -> {cpt_code: count}.
    - cpt_positions: Dict mapping CPT -> column index in the sheet.
    - is_merged: Function to resolve merged cell positions (row, col) -> (write_row, write_col).
    - output_filename: Filepath to save the updated workbook.
    """
    global Gross_encounters_col,sheet,week1_encounters_col_idx,week2_encounters_col_idx,weekly_counts # Uses global Gross_encounters_col

    if Gross_encounters_col is None:
        warning_msg = "Warning: 'Gross Encounters' column not found in payroll file. Gross Encounters will not be updated."
        logging.warning(warning_msg)
        processing_warnings.append(warning_msg)
    
    if week1_encounters_col_idx is None or week2_encounters_col_idx is None:
        warning_msg = "Warning: 'Week 1 Encounters' or 'Week 2 Encounters' columns not found in payroll file. Weekly encounters will not be updated."
        logging.warning(warning_msg)
        processing_warnings.append(warning_msg)
    output_full_path = os.path.join(get_base_dir(), output_filename)
    for name in common_providers:
        if name['payroll_name'] in practitioner_list:
            name_index = practitioner_list.index(name['payroll_name'])
            charge_capture_name = name['charge_capture_name']
            if name['charge_capture_name'] in provider_cpt_dict:
                cpt_found = False
                Gross_encounters = 0
                total_week1=0
                total_week2=0
                for cpt, col_idx in cpt_positions.items():
                    if cpt  in provider_cpt_dict[ name['charge_capture_name']]:
                        cpt_count = provider_cpt_dict[ name['charge_capture_name']][cpt]
                        filtered = weekly_counts[
                            (weekly_counts['Provider'] == charge_capture_name) &
                            (weekly_counts['CPT_Code'] == cpt)
                        ]
                        
                        # Get week1 and week2 counts
                        week1_count = filtered.loc[filtered['Week_Label'] == 'week1', 'Count'].sum()
                        week2_count = filtered.loc[filtered['Week_Label'] == 'week2', 'Count'].sum()
                        Gross_encounters+=cpt_count
                        total_week1+=week1_count
                        total_week2+=week2_count
                        # Get corrected cell position
                        row = name_index + 7
                        col = col_idx + 1
                        write_row, write_col = is_merged(row, col)
                        # print("✅ Matched CPT", cpt, "for", name, "at", row, write_col)
                        # Write only to top-left of merged range
                        sheet.cell(row=row, column=write_col).value = cpt_count
                        sheet.cell(row=row, column=write_col+3).value = week1_count
                        sheet.cell(row=row, column=write_col+4).value = week2_count
                        cpt_found = True
                    # Update Gross Encounters column
                    if Gross_encounters_col is not None:
                        sheet.cell(row=name_index + 7, column=Gross_encounters_col + 1).value = Gross_encounters
                    
                # Update Weekly Encounters columns
                if week1_encounters_col_idx is not None and week2_encounters_col_idx is not None:

                    sheet.cell(row=name_index + 7, column=week1_encounters_col_idx + 1).value =total_week1
                
                    sheet.cell(row=name_index + 7, column=week2_encounters_col_idx + 1).value =total_week2
                # if not cpt_found:
                #     print("❌ No valid CPT found for", name)
                #     print ("found intead",not_recognized[normalize_name])
            else:
                print(f"⚠️ Provider not found in dict: '{name}'")
    # current_workbook.save(output_filename)
def apply_manual_cpt_updates():
    global provider_cpt_dict, manual_cpt_updates
    logging.info(f"Applying {len(manual_cpt_updates)} manual CPT updates.")
    for update in manual_cpt_updates:
        provider = update["provider"]
        cpt = update["cpt"]
        update_cpt_counts(provider_cpt_dict, provider, cpt)
    logging.info("Manual CPT updates applied.")
#  process the sheet (payroll , and capture sheet), first function to init most of the global varibales 
def process_files(payroll_path, capture_path, payroll_sheet, output_filename, date_range_param):
    reset_global_variables()
    # Set the global date_range from parameter
    global date_range
    date_range = date_range_param  # Assign parameter to global variable
    # execel sheet
    global current_workbook,sheet
    # data structure
    global provider_cpt_dict,practitioner_list,cpt_positions,not_recognized,common_providers,weekly_counts
    # table boundaries
    global  start_row,row_end,merged_cells,Gross_encounters_col,week1_encounters_col_idx,week2_encounters_col_idx
    # orignal data frames
    global payroll_df, capture_df,cpt_counts
    # Load data
    # --- Load Payroll File ---
    try:
        payroll_df = pd.read_excel(payroll_path, sheet_name=payroll_sheet)
        logging.info(f"Successfully loaded payroll file: {payroll_path}, sheet: {payroll_sheet}, date range: {date_range_param}")
    except ValueError as e:
        msg = f"Invalid payroll sheet name or sheet not found: '{payroll_sheet}'. Please check the sheet name and try again. Details: {e}"
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise ValueError(msg)
    except FileNotFoundError:
        msg = f"Payroll file not found at {payroll_path}. Please ensure the file was uploaded correctly."
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise FileNotFoundError(msg)
    except Exception as e:
        msg = f"Failed to read payroll file '{payroll_path}'. Details: {e}"
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise Exception(msg)

    # --- Load Capture File ---
    try:
        capture_df = pd.read_excel(capture_path)
        logging.info(f"Successfully loaded capture file: {capture_path}")
    except FileNotFoundError:
        msg = f"Charge Capture file not found at {capture_path}. Please ensure the file was uploaded correctly."
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise FileNotFoundError(msg)
    except Exception as e:
        msg = f"Failed to read capture file '{capture_path}'. Details: {e}"
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise Exception(msg)

    # --- Validate essential columns in capture_df ---
    required_capture_cols = ["Provider", "CPT Codes", "DOS"]
    if not all(col in capture_df.columns for col in required_capture_cols):
        missing = [col for col in required_capture_cols if col not in capture_df.columns]
        msg = f"Charge Capture file is missing required columns: {', '.join(missing)}. Please ensure the file has 'Provider', 'CPT Codes', and 'DOS' columns."
        logging.error(msg)
        processing_warnings.append(f"Error: {msg}")
        raise ValueError(msg)

    # --- Load Excel Workbook for direct manipulation ---
    try:
        current_workbook = load_workbook(payroll_path)
        sheet = current_workbook[payroll_sheet]
        merged_cells = sheet.merged_cells.ranges
        logging.info(f"Successfully loaded workbook and selected sheet '{payroll_sheet}'.")
    except KeyError:
        msg = f"Sheet '{payroll_sheet}' not found in the payroll Excel file. Please enter the correct sheet name."
        logging.error(msg)
        processing_warnings.append(f"Error: {msg}")
        raise ValueError(msg)
    except Exception as e:
        msg = f"Failed to load payroll workbook or access sheet. Details: {e}"
        logging.error(msg, exc_info=True)
        processing_warnings.append(f"Error: {msg}")
        raise Exception(msg)
    global start_row
    try:
        start_row, cpt_positions = find_header_and_cpt_positions(payroll_df)
        logging.info(f"CPT headers found at row {start_row}. CPT positions: {cpt_positions}")
    except ValueError as e:
        logging.error(f"Error finding CPT headers in payroll file: {e}", exc_info=True)
        processing_warnings.append(f"Error: {str(e)}")
        raise e # Re-raise to be caught by the Flask route

    # Find the end of the data table using "Encounter" cell
    Encounter_cell_pos = find_cell_matches(payroll_df, "Encounter")
    if not Encounter_cell_pos:
        msg = "Could not find 'Encounter' cell in payroll file to determine the end of the data table. Please ensure this cell exists."
        logging.error(msg)
        processing_warnings.append(f"Error: {msg}")
        raise ValueError(msg)
    global row_end # Needed as it's modified
    r_span, c_span = get_cell_span(sheet, Encounter_cell_pos[0]['row']+2, Encounter_cell_pos[0]['col']+1)

    row_end=Encounter_cell_pos[0]['row']+r_span+1
    practitioner_list=extract_practitioners(payroll_df,start_row=start_row,stop_row=row_end)    
    Gross_encounters_dict=find_cell_matches(payroll_df, "Gross Encounters")
    Gross_encounters_col=Gross_encounters_dict[0]['col']+1
     # Find column indices for writing
    week1_encounters_col_match = find_cell_matches(payroll_df, "Week 1\nEncounters")
    week2_encounters_col_match = find_cell_matches(payroll_df, "Week 2\nEncounters")

    week1_encounters_col_idx = week1_encounters_col_match[0]['col'] if week1_encounters_col_match else None
    week2_encounters_col_idx = week2_encounters_col_match[0]['col'] if week2_encounters_col_match else None

    # Prepare raw name dictionaries
    Pay_roll_providers = {name: name for name in practitioner_list}
    charge_capture_providers = {name: name for name in capture_df["Provider"].unique()}

    try:
        weekly_counts = get_weekly_counts(capture_df.copy(), date_range_param)  # Pass a copy to avoid modifying original df
        if isinstance(weekly_counts, str):
        # Raise exception if function returned a message instead of a DataFrame
            raise ValueError(f"{weekly_counts} This might indicate that "\
                                    "no records in the Charge Capture file 'DOS' column fall within " \
                                   "the date range specified in the payroll sheet name"\
                                    " Please check the date period matching the Charge Capture DOSs.")

    except ValueError as e:
        logging.error(f"Error in weekly counts calculation: {e}")
        raise e  # Re-raise to be caught by the Flask route
    except TypeError as e:  # Catch TypeError if get_weekly_counts returned a non-DataFrame and it wasn't caught inside
        logging.critical(f"Critical Type Error: Weekly counts is not a DataFrame. {e}", exc_info=True)
        processing_warnings.append(f"Critical Internal Error: Weekly counts data is corrupted. Please report this. Details: {e}")
        weekly_counts = pd.DataFrame(columns=['Provider', 'Week_Label', 'CPT_Code', 'Count'])
        raise e
    except Exception as e:
        logging.error(f"An unexpected error occurred during weekly counts calculation: {e}")
        raise Exception(f"Failed to calculate weekly counts: {e}")
    # Initialize results
    common_providers = []
    only_in_Pay_roll_providers = []
    Gross_encounters_col = ((find_cell_matches(payroll_df, "Gross Encounters"))[0]['col'])
    match_results = match_providers(charge_capture_providers, Pay_roll_providers)
    cpt_counts = capture_df.groupby(['Provider', 'CPT Codes']).size().reset_index(name='Counts')
    if not provider_cpt_dict :
        process_cpt_counts(cpt_counts,cpt_positions)
    apply_manual_cpt_updates() 
    # write_provider_cpt_data_to_sheet(
    #     payroll_df=payroll_df,
    #     common_providers=common_providers,
    #     practitioner_list=practitioner_list,
    #     provider_cpt_dict=provider_cpt_dict,
    #     cpt_positions=cpt_positions,
    #     output_filename=output_filename
    # )
    
    print("done and saved")

    missing_providers = []
    if capture_df is not None and 'CPT Codes' in capture_df.columns and 'Provider' in capture_df.columns:
        for provider in capture_df[capture_df['CPT Codes'].isna()]['Provider'].unique():
            if pd.isna(provider):
                continue
            # Create mask for current provider with missing CPT codes
            mask = (capture_df['Provider'] == provider) & capture_df['CPT Codes'].isna()
            # Get positions and DOS values
            positions = capture_df.index[mask].tolist()
            dos_values = capture_df.loc[mask, 'DOS'].tolist() if 'DOS' in capture_df.columns else []
            # Append dictionary to result list
            missing_providers.append({
                'provider': str(provider),
                'positions': positions,
                'DOS': dos_values
            })
    
    return {
        'invalid_cpts': not_recognized,
        'missing_positions': capture_df[capture_df['CPT Codes'].isna()].index.tolist(),
        'missing_providers': missing_providers ,
    }

def get_weekly_counts(charge_capture_df: pd.DataFrame, date_range: str) -> pd.DataFrame | str:
    # Step 1: Parse the two dates
    start_str, end_str = date_range.split("-")
    start_date = pd.to_datetime(start_str.strip(), format='%m.%d.%y')
    end_date = pd.to_datetime(end_str.strip(), format='%m.%d.%y')

    # Step 2: Convert DOS to datetime
    charge_capture_df['DOS'] = pd.to_datetime(charge_capture_df['DOS'])

    # Step 3: Filter rows within the date range
    in_range_df = charge_capture_df[(charge_capture_df['DOS'] >= start_date) & (charge_capture_df['DOS'] <= end_date)]

    if in_range_df.empty:
        return "No records found in the selected date range."

    # Step 4: Define week boundaries
    week1_start = start_date
    week1_end = week1_start + timedelta(days=6)
    week2_start = week1_end + timedelta(days=1)
    week2_end = min(end_date, week2_start + timedelta(days=6))

    # Step 5: Assign Week_Label
    def label_week(d):
        if week1_start <= d <= week1_end:
            return 'week1'
        elif week2_start <= d <= week2_end:
            return 'week2'
        else:
            return None

    in_range_df['Week_Label'] = in_range_df['DOS'].apply(label_week)
    filtered_df = in_range_df[in_range_df['Week_Label'].notnull()]

    # Step 6: Group by Provider, Week, and CPT Code string (may contain multiple codes)
    grouped = filtered_df.groupby(['Provider', 'Week_Label', 'CPT Codes'])
    
    # Step 7: Expand multiple CPT Codes
    output = []

    for (provider, week, cpt_str), group in grouped:
        count = len(group)
        if pd.isna(cpt_str) or cpt_str.strip() == '':
            output.append({
                'Provider': provider,
                'Week_Label': week,
                'CPT_Code': None,
                'Count': count
            })
        else:
            codes = [code.strip() for code in cpt_str.split(",")]
            for code in codes:
                output.append({
                    'Provider': provider,
                    'Week_Label': week,
                    'CPT_Code': code,
                    'Count': count  # count stays the same for each code
                })

    result_df = pd.DataFrame(output)

    # Step 8: Aggregate again to sum counts for same Provider, Week_Label, CPT_Code
    final_df = result_df.groupby(['Provider', 'Week_Label', 'CPT_Code'], as_index=False)['Count'].sum()
    final_df = final_df.sort_values(by=['Provider', 'Week_Label', 'CPT_Code'])

    return final_df

def increment_encounter(provider: str, dos: str, weekly_counts_df: pd.DataFrame, date_range: str, cpt_code: str) -> pd.DataFrame:
    from datetime import datetime, timedelta
    import pandas as pd

    if isinstance(dos, str):
        date_obj = datetime.strptime(dos, "%m/%d/%Y")
    else:
        date_obj = dos

    # Parse the date range
    start_str, end_str = date_range.split("-")
    start_date = pd.to_datetime(start_str.strip(), format='%m.%d.%y')
    end_date = pd.to_datetime(end_str.strip(), format='%m.%d.%y')

    # Define week windows
    week1_start = start_date
    week1_end = week1_start + timedelta(days=6)
    week2_start = week1_end + timedelta(days=1)
    week2_end = min(end_date, week2_start + timedelta(days=6))

    # Assign week label
    if week1_start <= pd.Timestamp(date_obj) <= week1_end:
        week_label = 'week1'
    elif week2_start <= pd.Timestamp(date_obj) <= week2_end:
        week_label = 'week2'
    else:
        raise ValueError(f"DOS {date_obj.date()} is outside the date range: {date_range}")

    # Locate row based on provider, week, and CPT code
    mask = (
        (weekly_counts_df['Provider'] == provider) &
        (weekly_counts_df['Week_Label'] == week_label) &
        (weekly_counts_df['CPT_Code'] == cpt_code)
    )

    if not weekly_counts_df[mask].empty:
        weekly_counts_df.loc[mask, 'Count'] += 1
    else:
        new_row = {
            'Provider': provider,
            'Week_Label': week_label,
            'CPT_Code': cpt_code,
            'Count': 1
        }
        weekly_counts_df = pd.concat([weekly_counts_df, pd.DataFrame([new_row])], ignore_index=True)

    return weekly_counts_df.sort_values(by=['Provider', 'Week_Label', 'CPT_Code'])

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/process', methods=['POST'])
def handle_process():
    if 'payrollFile' not in request.files or 'captureFile' not in request.files:
        return jsonify({'error': 'Missing files'}), 400
    
    global output_filename
    payroll_file = request.files['payrollFile']
    capture_file = request.files['captureFile']
    payroll_sheet = request.form.get('payrollSheet')
    date_range = request.form.get('dateRange')
    
    if not payroll_sheet:
        return jsonify({'error': 'Payroll sheet name is required'}), 400
    
    if not date_range:
        return jsonify({'error': 'Date range is required (format: MM.DD.YY-MM.DD.YY)'}), 400
    
    # Validate date range format
    import re
    if not re.match(r'^\d{2}\.\d{2}\.\d{2}-\d{2}\.\d{2}\.\d{2}$', date_range):
        return jsonify({'error': 'Invalid date range format. Use MM.DD.YY-MM.DD.YY (e.g., 08.01.25-08.14.25)'}), 400
    
    output_filename = request.form.get('outputFileName', 'processed_payroll.xlsx')
    
    try:
        # Save uploaded files
        payroll_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(payroll_file.filename or 'payroll.xlsx'))
        capture_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(capture_file.filename or 'capture.xlsx'))
        payroll_file.save(payroll_path)
        capture_file.save(capture_path)

        results = process_files(payroll_path, capture_path, payroll_sheet, output_filename, date_range)
        
        # Validate results structure
        if not isinstance(results, dict):
            return jsonify({'error': 'Invalid processing results'}), 500
            
        invalid_cpts_array = []
        if 'invalid_cpts' in results and results['invalid_cpts']:
            invalid_cpts_array = [
                {"name": codes.get("name", provider), "found": {k: v for k, v in codes.items() if k != "name"}}
                for provider, codes in results['invalid_cpts'].items()
            ]

        return jsonify({
            'invalidCPTs': invalid_cpts_array,
            'missingPositions': results.get('missing_positions', []),
            'missingProviders': results.get('missing_providers', []),
        })
        
    except Exception as e:
        logging.error(f"Error in handle_process: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/add_missing_cpt', methods=['POST'])
def handle_add_cpt():
    data = request.get_json()
    provider = data.get('provider')
    cpt = data.get('cpt')
    dos = data.get('dos')  # make sure it's included
    global provider_cpt_dict, weekly_counts, date_range
    if not provider or not cpt:
        return jsonify({'error': 'Missing provider or CPT'}), 400
        
    try:
        # Check if provider exists in provider_cpt_dict
        if provider not in provider_cpt_dict:
            return jsonify({'error': f'Provider {provider} not found in processed data'}), 400

        print(f"Before update: {provider_cpt_dict.get(provider, {})}")
        # Update the in-memory dict directly
        update_cpt_counts(provider_cpt_dict, provider, cpt)
        # Update the weekly counts DataFrame if dos is provided
        if dos and weekly_counts is not None and not weekly_counts.empty:
            weekly_counts = increment_encounter(provider, dos, weekly_counts, date_range, cpt)
        # Record this manual update to re-apply later if provider_cpt_dict gets reset
        manual_cpt_updates.append({"provider": provider, "cpt": cpt})
        print(f"After update: {provider_cpt_dict.get(provider, {})}")
        return jsonify({'success': True})
        
    except Exception as e:
        logging.error(f"Error in handle_add_cpt: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/save_changes', methods=['POST'])
def handle_save_changes():
    global sheet, payroll_df, common_providers, practitioner_list, provider_cpt_dict, cpt_positions, current_workbook, output_filename
    
    if not current_workbook or not sheet:
        return jsonify({'error': 'No workbook loaded. Process files first.'}), 400
        
    if not output_filename:
        # Fallback filename if not set, though it should be set by /process
        output_filename = "processed_payroll_download.xlsx"
        logging.warning("Output filename was not set, using default for download.")

    try:
        # Ensure the sheet is up-to-date with the latest data from provider_cpt_dict
        if payroll_df is not None and common_providers and practitioner_list and provider_cpt_dict and cpt_positions:
            write_provider_cpt_data_to_sheet(
                payroll_df=payroll_df,
                common_providers=common_providers,
                practitioner_list=practitioner_list,
                provider_cpt_dict=provider_cpt_dict,
                cpt_positions=cpt_positions,
                output_filename=output_filename
            )
        
        # Save workbook to a BytesIO stream
        file_stream = BytesIO()
        current_workbook.save(file_stream)
        file_stream.seek(0)  # Rewind the stream to the beginning

        logging.info(f"Workbook prepared for download as {output_filename}")

        return send_file(
            file_stream,
            as_attachment=True,
            download_name=output_filename,  # Use the globally set output_filename
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"Error in /save_changes: {e}", exc_info=True)
        return jsonify({'error': f"Failed to prepare file for download: {str(e)}"}), 500

@app.route('/add_new_cpt_column', methods=['POST'])
def add_new_cpt_column():
    global cpt_positions, sheet, not_recognized, provider_cpt_dict, Gross_encounters_col, week2_encounters_col_idx, week1_encounters_col_idx, cpt_counts
    
    data = request.get_json()
    cpt_code = data.get('cpt', '').strip()

    if not cpt_code:
        return jsonify({'success': False, 'error': 'CPT code is required.'}), 400

    # Check if CPT code already exists
    if cpt_code in cpt_positions:
        return jsonify({'success': False, 'error': 'CPT code already exists.'}), 400

    # Check if we have the necessary global variables initialized
    if not cpt_positions or not sheet:
        return jsonify({'success': False, 'error': 'No payroll data loaded. Process files first.'}), 400

    try:
        # Get the last CPT position to calculate new position
        if not cpt_positions:
            return jsonify({'success': False, 'error': 'No existing CPT positions found.'}), 400
            
        last_cpt_index = cpt_positions[list(cpt_positions.keys())[-1]]
        new_cpt_index = last_cpt_index + 6  # not zero based 
        
        # Add the new CPT column to the sheet
        success = add_new_cpt(cpt_code, new_cpt_index)
        if not success:
            return jsonify({'success': False, 'error': 'Failed to add CPT column to sheet.'}), 500
            
        # Update positions and indices
        cpt_positions[cpt_code] = new_cpt_index - 1  # zero based 
        if Gross_encounters_col is not None:
            Gross_encounters_col = Gross_encounters_col + 5
        if week2_encounters_col_idx is not None:
            week2_encounters_col_idx = week2_encounters_col_idx + 5
        if week1_encounters_col_idx is not None:
            week1_encounters_col_idx = week1_encounters_col_idx + 5
        
        # Reprocess CPT counts with updated positions
        if cpt_counts is not None and not cpt_counts.empty:
            process_cpt_counts(cpt_counts, cpt_positions)
            apply_manual_cpt_updates()
            
            # Update the sheet with new data
            write_provider_cpt_data_to_sheet(
                payroll_df=payroll_df,
                common_providers=common_providers,
                practitioner_list=practitioner_list,
                provider_cpt_dict=provider_cpt_dict,
                cpt_positions=cpt_positions,
                output_filename=output_filename
            )
        
        return jsonify({'success': True, 'invalid_cpts': not_recognized}), 200
        
    except Exception as e:
        logging.error(f"Error in add_new_cpt_column: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
# --- Application Control (Shutdown and Browser Launch) ---

@app.route('/shutdown', methods=['POST'])
def shutdown():
    shutdown_server()
    return jsonify({"status": "Server shutting down..."})

def shutdown_server():
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        print("Warning: Not running with the Werkzeug Server.")
        os._exit(0)  # Forcefully terminate the Python process
    func()

# # Function to open the default web browser
# def open_browser():
#     webbrowser.open_new("http://localhost:5000")

# Main application entry point (commented out as this is imported by main.py)
# if __name__ == '__main__':
#     threading.Timer(1, open_browser).start()  # Open the browser after 1 second
#     app.run(host='localhost', port=5000)
